#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
マークダウンファイルからExcelファイルを作成するスクリプト（改良版）
すべてのセクションを含む包括的なExcelファイルを生成
"""

import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
import os

def parse_markdown_to_excel(md_file_path, excel_file_path):
    """マークダウンファイルを解析してExcelファイルを作成"""
    
    # マークダウンファイルを読み込み
    with open(md_file_path, 'r', encoding='utf-8') as file:
        content = file.read()
    
    # Excelワークブックを作成
    wb = Workbook()
    
    # 各シートを作成
    create_basic_info_sheet(wb, content)
    create_specialty_areas_sheet(wb, content)
    create_technical_skills_sheet(wb, content)
    create_self_pr_sheet(wb, content)
    create_project_experience_sheet(wb, content)
    create_responsibility_matrix_sheet(wb, content)
    create_strengths_sheet(wb, content)
    
    # デフォルトシートを削除
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Excelファイルを保存
    wb.save(excel_file_path)
    print(f"改良版Excelファイルが作成されました: {excel_file_path}")

def create_basic_info_sheet(wb, content):
    """基本情報シートを作成"""
    ws = wb.create_sheet(title="基本情報")
    
    # ヘッダースタイル
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # 基本情報を抽出
    basic_info = extract_basic_info(content)
    
    # データを追加
    ws['A1'] = "項目"
    ws['B1'] = "内容"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['B1'].font = header_font
    ws['B1'].fill = header_fill
    
    row = 2
    for key, value in basic_info.items():
        ws[f'A{row}'] = key
        ws[f'B{row}'] = value
        row += 1
    
    # 列幅を調整
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40

def create_specialty_areas_sheet(wb, content):
    """得意分野シートを作成"""
    ws = wb.create_sheet(title="得意分野")
    
    # ヘッダースタイル
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # 得意分野を抽出
    specialty_data = extract_specialty_areas(content)
    
    # データを追加
    ws['A1'] = "カテゴリ"
    ws['B1'] = "内容"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['B1'].font = header_font
    ws['B1'].fill = header_fill
    
    row = 2
    for category, items in specialty_data.items():
        if isinstance(items, list):
            for item in items:
                ws[f'A{row}'] = category
                ws[f'B{row}'] = item
                row += 1
        else:
            ws[f'A{row}'] = category
            ws[f'B{row}'] = items
            row += 1
    
    # 列幅を調整
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50

def create_technical_skills_sheet(wb, content):
    """技術スキルシートを作成"""
    ws = wb.create_sheet(title="技術スキル")
    
    # ヘッダースタイル
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # 技術スキルを抽出
    skills = extract_technical_skills(content)
    
    # データを追加
    ws['A1'] = "カテゴリ"
    ws['B1'] = "技術・言語"
    ws['C1'] = "経験年数"
    
    for col in ['A1', 'B1', 'C1']:
        ws[col].font = header_font
        ws[col].fill = header_fill
    
    row = 2
    for category, items in skills.items():
        for item, years in items.items():
            ws[f'A{row}'] = category
            ws[f'B{row}'] = item
            ws[f'C{row}'] = years
            row += 1
    
    # 列幅を調整
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15

def create_self_pr_sheet(wb, content):
    """自己PR・備考シートを作成"""
    ws = wb.create_sheet(title="自己PR・備考")
    
    # ヘッダースタイル
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # 自己PRを抽出
    self_pr_items = extract_self_pr(content)
    
    # データを追加
    ws['A1'] = "No"
    ws['B1'] = "自己PR・備考"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['B1'].font = header_font
    ws['B1'].fill = header_fill
    
    row = 2
    for i, item in enumerate(self_pr_items, 1):
        ws[f'A{row}'] = i
        ws[f'B{row}'] = item
        row += 1
    
    # 列幅を調整
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 80

def create_project_experience_sheet(wb, content):
    """プロジェクト経験シートを作成"""
    ws = wb.create_sheet(title="プロジェクト経験")
    
    # ヘッダースタイル
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # プロジェクト経験を抽出
    projects = extract_project_experience(content)
    
    # ヘッダーを設定
    headers = ["No", "会社名", "期間", "業種", "雇用形態", "チーム規模", "主要技術", "プロジェクト概要", "主な業務内容", "習得スキル", "成果・実績"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # データを追加
    for row, project in enumerate(projects, 2):
        ws.cell(row=row, column=1, value=project.get('no', ''))
        ws.cell(row=row, column=2, value=project.get('company', ''))
        ws.cell(row=row, column=3, value=project.get('period', ''))
        ws.cell(row=row, column=4, value=project.get('industry', ''))
        ws.cell(row=row, column=5, value=project.get('employment', ''))
        ws.cell(row=row, column=6, value=project.get('team_size', ''))
        ws.cell(row=row, column=7, value=project.get('technologies', ''))
        ws.cell(row=row, column=8, value=project.get('overview', ''))
        ws.cell(row=row, column=9, value=project.get('duties', ''))
        ws.cell(row=row, column=10, value=project.get('skills', ''))
        ws.cell(row=row, column=11, value=project.get('achievements', ''))
    
    # 列幅を調整
    column_widths = [5, 20, 20, 10, 12, 15, 30, 40, 40, 40, 40]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

def create_responsibility_matrix_sheet(wb, content):
    """担当領域シートを作成"""
    ws = wb.create_sheet(title="担当領域")
    
    # ヘッダースタイル
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # 担当領域マトリックスを抽出
    matrix_data = extract_responsibility_matrix(content)
    
    if matrix_data is not None and not matrix_data.empty:
        # DataFrameからExcelに変換
        for r in dataframe_to_rows(matrix_data, index=False, header=True):
            ws.append(r)
        
        # ヘッダーにスタイルを適用
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # 列幅を調整
        ws.column_dimensions['A'].width = 35
        for col in range(2, len(matrix_data.columns) + 1):
            ws.column_dimensions[chr(64 + col)].width = 18

def create_strengths_sheet(wb, content):
    """強み・特徴シートを作成"""
    ws = wb.create_sheet(title="強み・特徴")
    
    # ヘッダースタイル
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # 強み・特徴を抽出
    strengths = extract_strengths(content)
    
    # データを追加
    ws['A1'] = "No"
    ws['B1'] = "強み・特徴"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['B1'].font = header_font
    ws['B1'].fill = header_fill
    
    row = 2
    for i, strength in enumerate(strengths, 1):
        ws[f'A{row}'] = i
        ws[f'B{row}'] = strength
        row += 1
    
    # 列幅を調整
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 80

def extract_basic_info(content):
    """基本情報を抽出"""
    basic_info = {}
    
    # 基本情報テーブルを抽出
    table_match = re.search(r'## 📋 基本情報.*?\n((?:\|.*?\|.*?\n)+)', content, re.DOTALL)
    if table_match:
        table_content = table_match.group(1)
        for line in table_content.split('\n'):
            if '|' in line and '---' not in line and '項目' not in line:
                parts = [part.strip() for part in line.split('|') if part.strip()]
                if len(parts) >= 2:
                    key = parts[0].replace('**', '')
                    value = parts[1].replace('**', '')
                    basic_info[key] = value
    
    return basic_info

def extract_specialty_areas(content):
    """得意分野を抽出"""
    specialty_data = {}
    
    # 得意分野セクションを抽出
    specialty_match = re.search(r'## 🎯 得意分野(.*?)(?=## |$)', content, re.DOTALL)
    if specialty_match:
        specialty_content = specialty_match.group(1)
        
        # 得意分野リストを抽出（改良版）
        areas = []
        lines = specialty_content.split('\n')
        for line in lines:
            line = line.strip()
            if line.startswith('- **') and line.endswith('**'):
                area = line.replace('- **', '').replace('**', '')
                areas.append(area)
        if areas:
            specialty_data['得意分野'] = areas
        
        # 得意言語を抽出
        languages_match = re.search(r'### 得意言語(.*?)(?=### |$)', specialty_content, re.DOTALL)
        if languages_match:
            lang_text = languages_match.group(1)
            languages = re.findall(r'- (.+)', lang_text)
            specialty_data['得意言語'] = languages
        
        # 得意業務を抽出
        duties_match = re.search(r'### 得意業務(.*?)(?=### |$)', specialty_content, re.DOTALL)
        if duties_match:
            duties_text = duties_match.group(1)
            duties = re.findall(r'- (.+)', duties_text)
            specialty_data['得意業務'] = duties
    
    return specialty_data

def extract_technical_skills(content):
    """技術スキルを抽出"""
    skills = {
        "開発言語": {},
        "フレームワーク": {},
        "データベース": {},
        "サーバー・OS": {}
    }
    
    # 各カテゴリのテーブルを抽出
    patterns = {
        "開発言語": r'### 開発言語.*?\n((?:\|.*?\|.*?\n)+)',
        "フレームワーク": r'### フレームワーク.*?\n((?:\|.*?\|.*?\n)+)',
        "データベース": r'### データベース.*?\n((?:\|.*?\|.*?\n)+)',
        "サーバー・OS": r'### サーバー・OS.*?\n((?:\|.*?\|.*?\n)+)'
    }
    
    for category, pattern in patterns.items():
        match = re.search(pattern, content, re.DOTALL)
        if match:
            table_content = match.group(1)
            for line in table_content.split('\n'):
                if '|' in line and '---' not in line and '経験年数' not in line:
                    parts = [part.strip() for part in line.split('|') if part.strip()]
                    if len(parts) >= 2:
                        tech = parts[0]
                        years = parts[1]
                        skills[category][tech] = years
    
    return skills

def extract_self_pr(content):
    """自己PR・備考を抽出"""
    self_pr_items = []
    
    # 自己PR・備考セクションを抽出
    pr_match = re.search(r'## 🌟 自己PR・備考(.*?)(?=## |$)', content, re.DOTALL)
    if pr_match:
        pr_content = pr_match.group(1)
        # リスト項目を抽出
        items = re.findall(r'- (.+)', pr_content)
        self_pr_items = items
    
    return self_pr_items

def extract_project_experience(content):
    """プロジェクト経験を抽出"""
    projects = []
    
    # プロジェクトセクション全体を抽出
    project_section_match = re.search(r'## 📈 職歴・プロジェクト経験（時系列順）(.*?)(?=## 📊 担当領域|$)', content, re.DOTALL)
    if not project_section_match:
        return projects
    
    project_section_content = project_section_match.group(1)
    
    # プロジェクトを区切る---で分割
    project_sections = re.split(r'\n---\n', project_section_content)
    
    for section in project_sections:
        section = section.strip()
        if not section:
            continue
            
        # プロジェクトタイトルを抽出
        title_match = re.search(r'### (\d+)\. (.+?)（(.+?)）', section)
        if not title_match:
            continue
            
        project = {
            'no': title_match.group(1),
            'company': title_match.group(2),
            'period': title_match.group(3),
            'industry': '',
            'employment': '',
            'team_size': '',
            'technologies': '',
            'overview': '',
            'duties': '',
            'skills': '',
            'achievements': ''
        }
        
        # 業種、雇用形態、チーム規模を抽出
        info_match = re.search(r'\*\*期間：\*\*.*?\|\s*\*\*業種：\*\*\s*(.+?)\s*\|\s*\*\*雇用形態：\*\*\s*(.+?)\s*\n\*\*チーム規模：\*\*\s*(.+)', section)
        if info_match:
            project['industry'] = info_match.group(1).strip()
            project['employment'] = info_match.group(2).strip()
            project['team_size'] = info_match.group(3).strip()
        
        # 使用技術を抽出
        tech_match = re.search(r'#### 使用技術\n(.*?)(?=#### |$)', section, re.DOTALL)
        if tech_match:
            tech_content = tech_match.group(1).strip()
            tech_lines = []
            for line in tech_content.split('\n'):
                line = line.strip()
                if line.startswith('- **') and '：**' in line:
                    # - **言語・FW：** Python, Flask, React.js の形式
                    tech_line = line.replace('- **', '').replace('**', '')
                    tech_lines.append(tech_line)
            project['technologies'] = ' | '.join(tech_lines)
        
        # プロジェクト概要を抽出
        overview_match = re.search(r'#### プロジェクト概要\n(.*?)(?=#### |$)', section, re.DOTALL)
        if overview_match:
            overview_content = overview_match.group(1).strip()
            # 複数行を1つの文章に統合
            overview_content = re.sub(r'\n+', ' ', overview_content)
            project['overview'] = overview_content
        
        # 主な業務内容を抽出
        duties_match = re.search(r'#### 主な業務内容\n(.*?)(?=#### |$)', section, re.DOTALL)
        if duties_match:
            duties_content = duties_match.group(1).strip()
            duties_items = []
            lines = duties_content.split('\n')
            current_item = ""
            
            for line in lines:
                line = line.strip()
                if line.startswith('- **') and line.endswith('**'):
                    # - **業務内容** の形式
                    if current_item:
                        duties_items.append(current_item.strip())
                    current_item = line.replace('- **', '').replace('**', '')
                elif line.startswith('  - '):
                    # サブ項目
                    if current_item:
                        current_item += " " + line.replace('  - ', '')
                elif line.startswith('- ') and not line.startswith('  -'):
                    # 通常のリスト項目
                    if current_item:
                        duties_items.append(current_item.strip())
                    current_item = line.replace('- ', '')
                elif line and not line.startswith('-') and current_item:
                    # 継続行
                    current_item += " " + line
            
            if current_item:
                duties_items.append(current_item.strip())
            
            project['duties'] = ' | '.join(duties_items)
        
        # 習得スキルを抽出
        skills_match = re.search(r'#### 習得スキル\n(.*?)(?=#### |$)', section, re.DOTALL)
        if skills_match:
            skills_content = skills_match.group(1).strip()
            skills_items = []
            for line in skills_content.split('\n'):
                line = line.strip()
                if line.startswith('- ') and not line.startswith('  -'):
                    skill = line.replace('- ', '')
                    skills_items.append(skill)
            project['skills'] = ' | '.join(skills_items)
        
        # 成果・実績を抽出
        achievements_match = re.search(r'#### 成果・実績\n(.*?)(?=#### |$)', section, re.DOTALL)
        if achievements_match:
            achievements_content = achievements_match.group(1).strip()
            achievements_items = []
            for line in achievements_content.split('\n'):
                line = line.strip()
                if line.startswith('- ') and not line.startswith('  -'):
                    achievement = line.replace('- ', '')
                    achievements_items.append(achievement)
            project['achievements'] = ' | '.join(achievements_items)
        
        projects.append(project)
    
    return projects

def extract_responsibility_matrix(content):
    """担当領域マトリックスを抽出"""
    # 担当領域テーブルを探す
    matrix_match = re.search(r'## 📊 担当領域.*?\n((?:\|.*?\|.*?\n)+)', content, re.DOTALL)
    if not matrix_match:
        return None
    
    table_content = matrix_match.group(1)
    lines = [line.strip() for line in table_content.split('\n') if line.strip() and '---' not in line]
    
    if len(lines) < 2:
        return None
    
    # ヘッダーを抽出
    header_line = lines[0]
    headers = [col.strip() for col in header_line.split('|') if col.strip()]
    
    # データ行を抽出
    data_rows = []
    for line in lines[1:]:
        if '|' in line:
            row_data = [col.strip() for col in line.split('|') if col.strip()]
            if len(row_data) >= len(headers):
                data_rows.append(row_data[:len(headers)])
    
    if not data_rows:
        return None
    
    # DataFrameを作成
    df = pd.DataFrame(data_rows, columns=headers)
    return df

def extract_strengths(content):
    """強み・特徴を抽出"""
    strengths = []
    
    # 強み・特徴セクションを抽出
    strengths_match = re.search(r'## 🎯 強み・特徴(.*?)(?=## |$)', content, re.DOTALL)
    if strengths_match:
        strengths_content = strengths_match.group(1)
        # 番号付きリストを抽出
        items = re.findall(r'\d+\.\s+\*\*(.*?)\*\*:\s*(.*)', strengths_content)
        for item in items:
            strengths.append(f"{item[0]}: {item[1]}")
    
    return strengths

if __name__ == "__main__":
    # マークダウンファイルからExcelファイルを作成
    md_file = "HM_スキルシート.md"
    excel_file = "HM_スキルシート_improved.xlsx"
    
    if os.path.exists(md_file):
        parse_markdown_to_excel(md_file, excel_file)
    else:
        print(f"マークダウンファイルが見つかりません: {md_file}") 